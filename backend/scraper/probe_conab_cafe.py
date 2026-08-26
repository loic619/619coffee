"""
probe_conab_cafe.py — TEMPORARY. Locate CONAB's "Série Histórica do Café"
(production split arabica / conilon, by crop year) and dump its structure so a
parser can be written. Delete once the scraper exists.

Egress to conab.gov.br is blocked from the dev sandbox but open on CI, so this
has to run as a workflow.
"""
from __future__ import annotations

import io
import re
import sys

import requests

UA = {"User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                     "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")}
PAGES = [
    "https://portaldeinformacoes.conab.gov.br/download-arquivos.html",
    "https://portaldeinformacoes.conab.gov.br/safra-serie-historica-cafe.html",
]
# CONAB's file naming has drifted over the years; try the known shapes too.
GUESSES = [
    "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/SerieHistoricaCafe.xls",
    "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/SerieHistoricaCafe.xlsx",
    "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/SerieHistoricaCafe.txt",
]


def _dump_table(content: bytes, url: str) -> None:
    print(f"\n[file] {url}  {len(content):,} bytes  head={content[:8]!r}")
    # xlsx (zip) / xls (OLE) / csv-ish — try each without guessing from the name
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        print(f"[xlsx] sheets: {wb.sheetnames}")
        for name in wb.sheetnames[:4]:
            ws = wb[name]
            print(f"[xlsx] --- {name}: {ws.max_row} rows x {ws.max_column} cols")
            for i, row in enumerate(ws.iter_rows(max_row=14, values_only=True)):
                cells = [str(c)[:22] for c in row[:12] if c is not None]
                if cells:
                    print(f"[xlsx]   r{i}: {cells}")
        return
    except Exception as e:  # noqa: BLE001
        print(f"[xlsx] not xlsx ({type(e).__name__}: {str(e)[:80]})")
    try:
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        print(f"[xls] sheets: {book.sheet_names()}")
        for name in book.sheet_names()[:4]:
            sh = book.sheet_by_name(name)
            print(f"[xls] --- {name}: {sh.nrows} rows x {sh.ncols} cols")
            for i in range(min(14, sh.nrows)):
                cells = [str(c.value)[:22] for c in sh.row(i)[:12] if str(c.value).strip()]
                if cells:
                    print(f"[xls]   r{i}: {cells}")
        return
    except Exception as e:  # noqa: BLE001
        print(f"[xls] not xls ({type(e).__name__}: {str(e)[:80]})")
    txt = content[:1500].decode("latin-1", "replace")
    print(f"[text] {txt}")


def main() -> int:
    for page in PAGES:
        try:
            r = requests.get(page, headers=UA, timeout=30)
        except Exception as e:  # noqa: BLE001
            print(f"[page] {page} FAILED {e!r}")
            continue
        print(f"\n[page] {page} -> {r.status_code}, {len(r.text):,} chars")
        links = sorted(set(re.findall(r'href\s*=\s*["\']([^"\']+)["\']', r.text)))
        hits = [x for x in links if re.search(r"(cafe|café)", x, re.I)]
        print(f"[page] links mentioning cafe ({len(hits)}):")
        for h in hits[:25]:
            print(f"    {h}")
        files = [x for x in links if re.search(r"\.(xls|xlsx|csv|txt|zip)$", x, re.I)]
        print(f"[page] downloadable files ({len(files)}):")
        for f in files[:30]:
            print(f"    {f}")

    for url in GUESSES:
        try:
            r = requests.get(url, headers=UA, timeout=60)
        except Exception as e:  # noqa: BLE001
            print(f"\n[guess] {url} FAILED {e!r}")
            continue
        print(f"\n[guess] {url} -> {r.status_code} ({len(r.content):,} bytes)")
        if r.ok and len(r.content) > 2000:
            _dump_table(r.content, url)
            break
    return 0


if __name__ == "__main__":
    sys.exit(main())
